
# For complete honesty I made some of the data processing part and my friend Saif Ahmed made the streamlit UI but Gemini made the it part of mapping the pages
# Libraries and Gemini API configuration
import streamlit as st
from PIL import Image
import base64
import json
import re
from io import BytesIO
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
import google.generativeai as genai

# Gemini API key could replace later for any other agent if needed

API_KEY = "Your API key" # My own API key
genai.configure(api_key=API_KEY)

st.set_page_config(page_title="Production Report Scanner V14", layout="wide")


# Gemini prompt (made gemini write it to be honest)

GEMINI_PROMPT_PAGE1 = """
You are reading a handwritten production log form (Page 1 of 3).
The form contains:
- GENERAL INFO: Date (DD/MM/YYYY), Shift (Day/Night), Machine No., Production Line, Factory (Badr 1/Badr 2/Badr 3/Sadat)
- PRODUCT ENTRY #1: Shape (Hexagonal/Rectangular/I/Square 10x10/Square 20x20/Square 30x30/Square 40x40/Square 10x20/Square 20x40/Square 30x60/Square 60x60), Size (6cm/8cm), Granix (Yes/No), Strength (250/300/350/400/450/500/550), Color (Red/Yellow/Gray/Black/Beige/Dark Gray/Light Gray/Middle Gray), Pressed QTY
- ROUGH MIX QTY: Mix QTY, Cement 42.5, Cement 52.5, White Cement, Sand, Adsa 0, Sen 1, Oxide Qty + color (Red/Yellow/Black)
- FACE MIX QTY: Mix QTY, Cement 42.5, Cement 52.5, White Cement, Sand, Oxide Qty + color (Red/Yellow/Black)

Extract ALL filled-in values (numbers in boxes, checked checkboxes).
For checkboxes: a box is checked if it contains any mark (X, tick, cross, scribble, filled, circled, slash).
For numbers: read digits written in individual small boxes left to right; ignore empty boxes.

Return ONLY valid JSON, no explanation, no markdown:
{
  "Date": "DD/MM/YYYY or empty string",
  "Shift": "Day or Night or empty string",
  "Machine": "number string or empty string",
  "Production_line": "value or empty string",
  "Factory": "Badr 1 or Badr 2 or Badr 3 or Sadat or empty string",
  "Shape_1": "shape string or empty string",
  "Size_1": "6cm or 8cm or empty string",
  "Granix_1": "Yes or No or empty string",
  "Strength_1": "number string or empty string",
  "Color_1": "color string or empty string",
  "QTY_1": "number string or empty string",
  "Rough_Mix_QTY_1": "number string or empty string",
  "Rough_Cement_42_1": "number string or empty string",
  "Rough_Cement_52_1": "number string or empty string",
  "Rough_White_Cement_1": "number string or empty string",
  "Rough_Sand_1": "number string or empty string",
  "Rough_Adsa_1": "number string or empty string",
  "Rough_Sen1_1": "number string or empty string",
  "Rough_Oxide_Qty_1": "number string or empty string",
  "Rough_Oxide_Color_1": "Red or Yellow or Black or empty string",
  "Face_Mix_QTY_1": "number string or empty string",
  "Face_Cement_42_1": "number string or empty string",
  "Face_Cement_52_1": "number string or empty string",
  "Face_White_Cement_1": "number string or empty string",
  "Face_Sand_1": "number string or empty string",
  "Face_Oxide_Qty_1": "number string or empty string",
  "Face_Oxide_Color_1": "Red or Yellow or Black or empty string"
}
"""

GEMINI_PROMPT_PAGE2 = """
You are reading a handwritten production log form (Page 2 of 3: PRODUCT ENTRY #2).
This page is identical in structure to page 1's product section but for the second product.
Return ONLY valid JSON:
{
  "Shape_2": "shape string or empty string",
  "Size_2": "6cm or 8cm or empty string",
  "Granix_2": "Yes or No or empty string",
  "Strength_2": "number string or empty string",
  "Color_2": "color string or empty string",
  "QTY_2": "number string or empty string",
  "Rough_Mix_QTY_2": "number string or empty string",
  "Rough_Cement_42_2": "number string or empty string",
  "Rough_Cement_52_2": "number string or empty string",
  "Rough_White_Cement_2": "number string or empty string",
  "Rough_Sand_2": "number string or empty string",
  "Rough_Adsa_2": "number string or empty string",
  "Rough_Sen1_2": "number string or empty string",
  "Rough_Oxide_Qty_2": "number string or empty string",
  "Rough_Oxide_Color_2": "Red or Yellow or Black or empty string",
  "Face_Mix_QTY_2": "number string or empty string",
  "Face_Cement_42_2": "number string or empty string",
  "Face_Cement_52_2": "number string or empty string",
  "Face_White_Cement_2": "number string or empty string",
  "Face_Sand_2": "number string or empty string",
  "Face_Oxide_Qty_2": "number string or empty string",
  "Face_Oxide_Color_2": "Red or Yellow or Black or empty string"
}
"""

GEMINI_PROMPT_PAGE3 = """
You are reading a handwritten production log form (Page 3 of 3: MAINTENANCE & DOWNTIME).
Return ONLY valid JSON. Times as "HH:MM" strings or empty string:
{
  "issues": [
    {"type": "Electrical or Mechanical or empty string", "from": "HH:MM or empty", "to": "HH:MM or empty"},
    {"type": "Electrical or Mechanical or empty string", "from": "HH:MM or empty", "to": "HH:MM or empty"},
    {"type": "Electrical or Mechanical or empty string", "from": "HH:MM or empty", "to": "HH:MM or empty"},
    {"type": "Electrical or Mechanical or empty string", "from": "HH:MM or empty", "to": "HH:MM or empty"}
  ],
  "tools": [
    {"tool": "Loader or Clark or Solar or empty string", "from": "HH:MM or empty", "to": "HH:MM or empty"},
    {"tool": "Loader or Clark or Solar or empty string", "from": "HH:MM or empty", "to": "HH:MM or empty"}
  ],
  "missing": {
    "Cement 42.5":  {"checked": true, "from": "HH:MM", "to": "HH:MM"},
    "Cement 52.5":  {"checked": false, "from": "", "to": ""},
    "White Cement": {"checked": false, "from": "", "to": ""},
    "Sand":         {"checked": false, "from": "", "to": ""},
    "Adsa 0":       {"checked": false, "from": "", "to": ""},
    "Sen 1":        {"checked": false, "from": "", "to": ""},
    "Oxide Red":    {"checked": false, "from": "", "to": ""},
    "Oxide Yellow": {"checked": false, "from": "", "to": ""},
    "Oxide Black":  {"checked": false, "from": "", "to": ""}
  },
  "signoff": {
    "supervisor": "name or empty string",
    "mixer_operator": "name or empty string",
    "machine_operator": "name or empty string",
    "forklift_operator": "name or empty string",
    "verified_by": "name or empty string"
  }
}
"""

# gemini interaction and mapping functions to help it extract the data in a structured way, also added some error handling to avoid crashes and get insights on what went wrong in case of issues with gemini response or parsing

# Converts a PIL image to Gemini's expected inline data format (base64-encoded JPEG)

def _pil_to_gemini_part(image: "Image.Image"):
    buf = BytesIO()
    image.save(buf, format="JPEG", quality=95)
    return {"mime_type": "image/jpeg", "data": base64.b64encode(buf.getvalue()).decode("utf-8")}

# Calls Gemini API with the given image and prompt, returns parsed JSON or empty dict on failure.
def _call_gemini(image: "Image.Image", prompt: str) -> dict:
    model = genai.GenerativeModel("gemini-3-flash-preview")
    response = model.generate_content([
        {"role": "user", "parts": [{"inline_data": _pil_to_gemini_part(image)}, {"text": prompt}]}
    ])
    raw = response.text.strip().replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"❌ Gemini JSON parse error: {e}\nRaw: {raw[:300]}")
        return {}

# Mapping functions to extract relevant fields from Gemini's response into our structured data format

def _map_gemini_page1(g: dict, data: dict):
    def s(k): return str(g.get(k, "") or "").strip()
    data["Date"], data["Shift"], data["Machine"] = s("Date"), s("Shift"), s("Machine")
    data["Production line"], data["Factory"] = s("Production_line"), s("Factory")
    
    data["Shape_1"], data["Size_1"], data["Granix_1"] = s("Shape_1"), s("Size_1"), s("Granix_1")
    data["Strength_1"], data["Color_1"], data["QTY_1"] = s("Strength_1"), s("Color_1"), s("QTY_1")
    
    data["Rough mix no._1"] = s("Rough_Mix_QTY_1")
    data["Cement 42.5 (Rough)_1"] = s("Rough_Cement_42_1")
    data["Cement 52.5 (Rough)_1"] = s("Rough_Cement_52_1")
    data["White Cement (Rough)_1"] = s("Rough_White_Cement_1")
    data["Sand (Rough)_1"] = s("Rough_Sand_1")
    data["Adsa 0 (Rough)_1"] = s("Rough_Adsa_1")
    data["Sen 1 (Rough)_1"] = s("Rough_Sen1_1")
    data["Oxide (Rough)_1"] = s("Rough_Oxide_Qty_1")
    data["Oxide (Rough) Color_1"] = s("Rough_Oxide_Color_1")
    
    data["Face mix no._1"] = s("Face_Mix_QTY_1")
    data["Cement 42.5 (Face)_1"] = s("Face_Cement_42_1")
    data["Cement 52.5 (Face)_1"] = s("Face_Cement_52_1")
    data["White Cement (Face)_1"] = s("Face_White_Cement_1")
    data["Sand (Face)_1"] = s("Face_Sand_1")
    data["Oxide (Face)_1"] = s("Face_Oxide_Qty_1")
    data["Oxide (Face) Color_1"] = s("Face_Oxide_Color_1")

# Page 2 has the same structure as page 1's product section but for the second product entry, so we can reuse the same mapping logic with different keys to avoid code duplication and keep it organized

def _map_gemini_page2(g: dict, data: dict):
    def s(k): return str(g.get(k, "") or "").strip()
    data["Shape_2"], data["Size_2"], data["Granix_2"] = s("Shape_2"), s("Size_2"), s("Granix_2")
    data["Strength_2"], data["Color_2"], data["QTY_2"] = s("Strength_2"), s("Color_2"), s("QTY_2")
    
    data["Rough mix no._2"] = s("Rough_Mix_QTY_2")
    data["Cement 42.5 (Rough)_2"] = s("Rough_Cement_42_2")
    data["Cement 52.5 (Rough)_2"] = s("Rough_Cement_52_2")
    data["White Cement (Rough)_2"] = s("Rough_White_Cement_2")
    data["Sand (Rough)_2"] = s("Rough_Sand_2")
    data["Adsa 0 (Rough)_2"] = s("Rough_Adsa_2")
    data["Sen 1 (Rough)_2"] = s("Rough_Sen1_2")
    data["Oxide (Rough)_2"] = s("Rough_Oxide_Qty_2")
    data["Oxide (Rough) Color_2"] = s("Rough_Oxide_Color_2")
    
    data["Face mix no._2"] = s("Face_Mix_QTY_2")
    data["Cement 42.5 (Face)_2"] = s("Face_Cement_42_2")
    data["Cement 52.5 (Face)_2"] = s("Face_Cement_52_2")
    data["White Cement (Face)_2"] = s("Face_White_Cement_2")
    data["Sand (Face)_2"] = s("Face_Sand_2")
    data["Oxide (Face)_2"] = s("Face_Oxide_Qty_2")
    data["Oxide (Face) Color_2"] = s("Face_Oxide_Color_2")

# Page 3 contains different types of data (maintenance issues, tool downtime, missing materials, and sign-off info), so we need a more complex mapping function to handle the nested structures and multiple entries for issues and tools. We also added error handling to avoid crashes if the expected keys are missing or have unexpected formats in Gemini's response.    

def _map_gemini_page3(g: dict, data: dict):
    issues = g.get("issues", [])
    for i in range(min(4, len(issues))):
        row = issues[i]
        kind = str(row.get("type", "")).strip().lower()
        frm, to = str(row.get("from", "")).strip(), str(row.get("to", "")).strip()
        if "electrical" in kind:
            data["Maintenance Electricity"][i], data["Maintenance Electricity From"][i], data["Maintenance Electricity To"][i] = True, frm, to
        elif "mechanical" in kind:
            data["Maintenance Mechanical"][i], data["Maintenance Mechanical From"][i], data["Maintenance Mechanical To"][i] = True, frm, to

    tools = g.get("tools", [])
    for i in range(min(2, len(tools))):
        row = tools[i]
        tool = str(row.get("tool", "")).strip().lower()
        frm, to = str(row.get("from", "")).strip(), str(row.get("to", "")).strip()
        if "loader" in tool:
            data["Tools Loader"][i], data["Tools Loader From"][i], data["Tools Loader To"][i] = True, frm, to
        elif "clark" in tool:
            data["Tools Clark"][i], data["Tools Clark From"][i], data["Tools Clark To"][i] = True, frm, to
        elif "solar" in tool:
            data["Tools Solar"][i], data["Tools Solar From"][i], data["Tools Solar To"][i] = True, frm, to

    missing = g.get("missing", {})
    mat_map = {
        "Cement 42.5": "Missing Cement 42.5", "Cement 52.5": "Missing Cement 52.5",
        "White Cement": "Missing White Cement", "Sand": "Missing Sand",
        "Adsa 0": "Missing Adsa 0", "Sen 1": "Missing Sen 1",
        "Oxide Red": "Missing Oxide Red", "Oxide Yellow": "Missing Oxide Yellow", "Oxide Black": "Missing Oxide Black",
    }
    for mat_key, data_key in mat_map.items():
        row = missing.get(mat_key, {})
        if isinstance(row, dict) and row.get("checked"):
            data[data_key] = True
            data[f"{data_key} From"] = str(row.get("from", ""))
            data[f"{data_key} To"] = str(row.get("to", ""))

    signoff = g.get("signoff", {})
    data["Supervisor Name"] = str(signoff.get("supervisor", ""))
    data["Mixer Operator"] = str(signoff.get("mixer_operator", ""))
    data["Machine Operator"] = str(signoff.get("machine_operator", ""))
    data["Forklift Operator"] = str(signoff.get("forklift_operator", ""))
    data["Verified By"] = str(signoff.get("verified_by", ""))

# Main function to scan all pages with Gemini and aggregate the extracted data into a single structured dictionary. It also collects debug messages to help identify any issues during the scanning and parsing process. We limit the number of pages to 3 since that's the expected number of pages in the form, and we use the corresponding prompt and mapping function for each page to ensure we extract the correct fields.

def scan_all_pages_gemini(images: list) -> tuple:
    data = {
        "Date": "", "Factory": "", "Machine": "", "Shift": "", "Production line": "",
        "Shape_1": "", "Size_1": "", "Granix_1": "", "Strength_1": "", "Color_1": "", "QTY_1": "",
        "Rough mix no._1": "", "Cement 42.5 (Rough)_1": "", "Cement 52.5 (Rough)_1": "", "White Cement (Rough)_1": "", "Adsa 0 (Rough)_1": "", "Sen 1 (Rough)_1": "", "Sand (Rough)_1": "", "Oxide (Rough)_1": "", "Oxide (Rough) Color_1": "",
        "Face mix no._1": "", "Cement 42.5 (Face)_1": "", "Cement 52.5 (Face)_1": "", "White Cement (Face)_1": "", "Sand (Face)_1": "", "Oxide (Face)_1": "", "Oxide (Face) Color_1": "",
        "Shape_2": "", "Size_2": "", "Granix_2": "", "Strength_2": "", "Color_2": "", "QTY_2": "",
        "Rough mix no._2": "", "Cement 42.5 (Rough)_2": "", "Cement 52.5 (Rough)_2": "", "White Cement (Rough)_2": "", "Adsa 0 (Rough)_2": "", "Sen 1 (Rough)_2": "", "Sand (Rough)_2": "", "Oxide (Rough)_2": "", "Oxide (Rough) Color_2": "",
        "Face mix no._2": "", "Cement 42.5 (Face)_2": "", "Cement 52.5 (Face)_2": "", "White Cement (Face)_2": "", "Sand (Face)_2": "", "Oxide (Face)_2": "", "Oxide (Face) Color_2": "",
        "Maintenance Electricity": [False]*4, "Maintenance Electricity From": [""]*4, "Maintenance Electricity To": [""]*4,
        "Maintenance Mechanical": [False]*4, "Maintenance Mechanical From": [""]*4, "Maintenance Mechanical To": [""]*4,
        "Tools Loader": [False]*2, "Tools Loader From": [""]*2, "Tools Loader To": [""]*2,
        "Tools Clark": [False]*2, "Tools Clark From": [""]*2, "Tools Clark To": [""]*2,
        "Tools Solar": [False]*2, "Tools Solar From": [""]*2, "Tools Solar To": [""]*2,
    }
    
    # Initialize Missing Materials
    for m in ["Cement 42.5", "Cement 52.5", "White Cement", "Sand", "Adsa 0", "Sen 1", "Oxide Red", "Oxide Yellow", "Oxide Black"]:
        data[f"Missing {m}"] = False
        data[f"Missing {m} From"] = ""
        data[f"Missing {m} To"] = ""
        
    for k in ["Supervisor Name", "Mixer Operator", "Machine Operator", "Forklift Operator", "Verified By"]:
        data[k] = ""

    debug = []
    prompts = [GEMINI_PROMPT_PAGE1, GEMINI_PROMPT_PAGE2, GEMINI_PROMPT_PAGE3]
    mappers = [_map_gemini_page1, _map_gemini_page2, _map_gemini_page3]

    for idx, image in enumerate(images[:3]):
        try:
            g = _call_gemini(image, prompts[idx])
            mappers[idx](g, data)
            debug.append(f"Page {idx+1} Parsed Successfully")
        except Exception as e:
            debug.append(f"Page {idx+1} Error: {e}")

    return data, debug

# Functions for data processing and Excel export

# Combines a list of PIL images vertically into a single image, centering them horizontally. This is useful for displaying all scanned pages together or for creating a single image to send to Gemini if needed. It handles cases where the list might be empty and ensures the combined image has a white background.

def combine_images_vertically(images):
    if not images: return None
    widths, heights = [img.width for img in images], [img.height for img in images]
    combined = Image.new('RGB', (max(widths), sum(heights)), color='white')
    y_offset = 0
    for img in images:
        combined.paste(img, ((max(widths) - img.width) // 2, y_offset))
        y_offset += img.height
    return combined

# Utility function to convert extracted string values to numbers, handling cases where the value might be empty or contain non-numeric characters. It uses regular expressions to clean the input and attempts to convert it to a float, returning 0 if conversion fails. This is important for the calculations we need to perform later on the extracted data.

def convert_to_number(value):
    if value is None or value == '': return 0
    try:
        clean = re.sub(r'[^\d.]', '', str(value))
        return float(clean) if clean else 0
    except: return 0

# Utility function to safely get the index of a value in a list of options, returning 0 if the value is not found. This can be useful for mapping extracted string values to predefined categories or for handling cases where the expected value might be missing or unrecognized without causing errors in the code.

def safe_index(options_list, value):
    return options_list.index(value) if value in options_list else 0

# Function to calculate the total quantities of materials used in the rough and face mixes for a given product entry. It takes the extracted data and the entry number (1 or 2) as input, retrieves the relevant mix quantities and material amounts, and performs the necessary calculations to return a dictionary with the calculated totals for each material in both tons (for cements) and kilograms (for sand, oxide, Adsa 0, and Sen 1). It also handles cases where the input values might be missing or non-numeric by using the convert_to_number function.

def calculate_mix_info(data, entry_num=1):
    calc = {}
    r_mix = convert_to_number(data.get(f'Rough mix no._{entry_num}'))
    f_mix = convert_to_number(data.get(f'Face mix no._{entry_num}'))
    
    # CEMENTS: (Rough + Face) / 1000 -> TONS
    for mat in ['Cement 42.5', 'Cement 52.5', 'White Cement']:
        r_val = convert_to_number(data.get(f'{mat} (Rough)_{entry_num}'))
        f_val = convert_to_number(data.get(f'{mat} (Face)_{entry_num}'))
        calc[f'{mat} (Calculated)_{entry_num}'] = round(((r_mix * r_val) + (f_mix * f_val)) / 1000, 3)
    
    # SAND & OXIDE: (Rough + Face) -> Standard units (Kg)
    for mat in ['Sand', 'Oxide']:
        r_val = convert_to_number(data.get(f'{mat} (Rough)_{entry_num}'))
        f_val = convert_to_number(data.get(f'{mat} (Face)_{entry_num}'))
        calc[f'{mat} (Calculated)_{entry_num}'] = round((r_mix * r_val) + (f_mix * f_val), 2)
    
    # ADSA 0 & SEN 1: Rough Only -> Standard units (Kg)
    for mat in ['Adsa 0', 'Sen 1']:
        r_val = convert_to_number(data.get(f'{mat} (Rough)_{entry_num}'))
        calc[f'{mat} (Calculated)_{entry_num}'] = round(r_mix * r_val, 2)
    
    return calc

# Function to export the extracted and calculated data into a well-formatted Excel file. It creates a new workbook, sets up the worksheet with appropriate column widths, styles for headers and subheaders, and populates the cells with the provided data. It also handles the inclusion of an image (if provided) in a separate sheet. The function is designed to be robust and can be extended to include additional formatting or error handling as needed.

def export_to_excel(data, calc_1, calc_2, image_bytes=None):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Production Report"
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 35
        h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        s_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        h_font, b_font = Font(bold=True, color="FFFFFF"), Font(bold=True)
        
        row = 1
        ws[f'A{row}'] = "DAILY PRODUCTION REPORT"; ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        def write_header(title):
            nonlocal row
            ws[f'A{row}'] = title; ws[f'A{row}'].fill, ws[f'A{row}'].font = h_fill, h_font; row += 1

        write_header("GENERAL INFORMATION")
        for field in ['Date', 'Factory', 'Machine', 'Shift', 'Production line']:
            ws[f'A{row}'], ws[f'B{row}'] = field, data.get(field, ''); ws[f'A{row}'].font = b_font; row += 1
        row += 1
        
        for p_num, calc_data in [(1, calc_1), (2, calc_2)]:
            write_header(f"PRODUCT ENTRY #{p_num}")
            for field in [f'Shape_{p_num}', f'Size_{p_num}', f'Color_{p_num}', f'Strength_{p_num}', f'Granix_{p_num}', f'QTY_{p_num}']:
                ws[f'A{row}'], ws[f'B{row}'] = field.replace(f'_{p_num}', ''), data.get(field, ''); ws[f'A{row}'].font = b_font; row += 1
            
            ws[f'A{row}'] = f"Calculated Material Totals #{p_num}"
            ws[f'A{row}'].fill, ws[f'A{row}'].font = s_fill, b_font; row += 1
            for mat in ['Cement 42.5', 'Cement 52.5', 'White Cement']:
                ws[f'A{row}'], ws[f'B{row}'] = f"{mat} (Tons)", calc_data.get(f'{mat} (Calculated)_{p_num}', 0); row += 1
            for mat in ['Sand', 'Oxide', 'Adsa 0', 'Sen 1']:
                ws[f'A{row}'], ws[f'B{row}'] = f"{mat} (Kg)", calc_data.get(f'{mat} (Calculated)_{p_num}', 0); row += 1
            row += 1

        write_header("MAINTENANCE & DOWNTIME")
        for i in range(4):
            if data["Maintenance Electricity"][i]:
                ws[f'A{row}'], ws[f'B{row}'] = f"Electrical #{i+1}", f"{data['Maintenance Electricity From'][i]} to {data['Maintenance Electricity To'][i]}"; row += 1
            if data["Maintenance Mechanical"][i]:
                ws[f'A{row}'], ws[f'B{row}'] = f"Mechanical #{i+1}", f"{data['Maintenance Mechanical From'][i]} to {data['Maintenance Mechanical To'][i]}"; row += 1
        row += 1

        write_header("TOOLS DOWNTIME")
        for tool in ['Loader', 'Clark', 'Solar']:
            for i in range(2):
                if data[f"Tools {tool}"][i]:
                    ws[f'A{row}'], ws[f'B{row}'] = f"{tool} #{i+1}", f"{data[f'Tools {tool} From'][i]} to {data[f'Tools {tool} To'][i]}"; row += 1
        row += 1

        write_header("MISSING MATERIAL")
        for m in ["Cement 42.5", "Cement 52.5", "White Cement", "Sand", "Adsa 0", "Sen 1", "Oxide Red", "Oxide Yellow", "Oxide Black"]:
            if data[f"Missing {m}"]:
                ws[f'A{row}'], ws[f'B{row}'] = m, f"{data[f'Missing {m} From']} to {data[f'Missing {m} To']}"; row += 1
        row += 1

        write_header("SIGN-OFF")
        for field in ['Supervisor Name', 'Mixer Operator', 'Machine Operator', 'Forklift Operator', 'Verified By']:
            ws[f'A{row}'], ws[f'B{row}'] = field, data.get(field, ''); ws[f'A{row}'].font = b_font; row += 1

        if image_bytes:
            try:
                ws2 = wb.create_sheet(title="Scanned Image")
                img = ExcelImage(BytesIO(image_bytes))
                img.width, img.height = 500, 650
                ws2.add_image(img, 'A1')
            except: pass
        
        temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp.name); temp.close()
        return temp.name
    except Exception as e:
        st.error(f"Excel Error: {e}"); return None

# Streamlit UI one of my friends made it all credits to Saif Ahmed :D

st.title("🏭 Production Report Scanner")

tab1, tab2, tab3 = st.tabs(["📤 Upload & Scan", "📊 Preview & Edit Data", "💾 Export"])

if 'extracted_data' not in st.session_state: st.session_state.extracted_data = None
if 'combined_image_bytes' not in st.session_state: st.session_state.combined_image_bytes = None

with tab1:
    st.info("📋 **Upload all 3 pages in order (Page 1, Page 2, Page 3)**")
    uploaded_files = st.file_uploader("Choose images", type=['png', 'jpg', 'jpeg'], accept_multiple_files=True)
    if uploaded_files:
        images = [Image.open(f) for f in uploaded_files]
        cols = st.columns(min(3, len(images)))
        for idx, img in enumerate(images): cols[idx % 3].image(img, width=300)
        
        buf = BytesIO(); combine_images_vertically(images).save(buf, format='PNG')
        st.session_state.combined_image_bytes = buf.getvalue()
        
        if st.button("🔍 Scan with Gemini", type="primary"):
            with st.spinner("Analyzing pages..."):
                data, debug = scan_all_pages_gemini(images)
                st.session_state.extracted_data = data
                st.success("✅ Scan complete! Move to the Preview tab.")

with tab2:
    st.header("Review & Edit Extracted Data")
    if st.session_state.extracted_data:
        d = st.session_state.extracted_data
        
        st.subheader("📋 GENERAL INFORMATION")
        c1, c2, c3 = st.columns(3)
        d['Date'] = c1.text_input("Date", d.get('Date', ''), key="gi_date")
        d['Factory'] = c1.text_input("Factory", d.get('Factory', ''), key="gi_factory")
        d['Machine'] = c2.text_input("Machine", d.get('Machine', ''), key="gi_machine")
        d['Shift'] = c2.selectbox("Shift", ['', 'Day', 'Night'], index=safe_index(['', 'Day', 'Night'], d.get('Shift', '')), key="gi_shift")
        d['Production line'] = c3.text_input("Production Line", d.get('Production line', ''), key="gi_prodline")

        def render_product_editor(entry_num):
            n = entry_num  # short alias for key strings
            st.subheader(f"🎨 PRODUCT ENTRY #{entry_num}")
            c1, c2, c3 = st.columns(3)
            d[f'Shape_{n}'] = c1.selectbox(f"Shape {n}", ['', 'Hexagonal', 'Rectangular', 'I', 'Square'],
                index=safe_index(['', 'Hexagonal', 'Rectangular', 'I', 'Square'], d.get(f'Shape_{n}', '')),
                key=f"shape_{n}")
            d[f'Size_{n}'] = c1.selectbox(f"Size {n}", ['', '8cm', '6cm'],
                index=safe_index(['', '8cm', '6cm'], d.get(f'Size_{n}', '')),
                key=f"size_{n}")
            d[f'Color_{n}'] = c2.selectbox(f"Color {n}", ['', 'Red', 'Yellow', 'Gray', 'Black', 'Beige'],
                index=safe_index(['', 'Red', 'Yellow', 'Gray', 'Black', 'Beige'], d.get(f'Color_{n}', '')),
                key=f"color_{n}")
            d[f'Strength_{n}'] = c2.text_input(f"Strength {n} (MPa)", d.get(f'Strength_{n}', ''),
                key=f"strength_{n}")
            d[f'QTY_{n}'] = c3.text_input(f"QTY {n} (pcs)", d.get(f'QTY_{n}', ''),
                key=f"qty_{n}")
            d[f'Granix_{n}'] = c3.selectbox(f"Granix {n}", ['', 'Yes', 'No'],
                index=safe_index(['', 'Yes', 'No'], d.get(f'Granix_{n}', '')),
                key=f"granix_{n}")

            with st.expander(f"🛠️ Edit Raw Mix Variables (Product {n})"):
                mc1, mc2 = st.columns(2)
                d[f'Rough mix no._{n}'] = mc1.text_input("Rough Mix QTY", d.get(f'Rough mix no._{n}', ''),
                    key=f"rough_mix_qty_{n}")
                d[f'Face mix no._{n}'] = mc2.text_input("Face Mix QTY", d.get(f'Face mix no._{n}', ''),
                    key=f"face_mix_qty_{n}")

                for mat in ['Cement 42.5', 'Cement 52.5', 'White Cement', 'Sand', 'Oxide']:
                    mat_key = mat.replace(' ', '_').replace('.', '')
                    rc1, rc2 = st.columns(2)
                    d[f'{mat} (Rough)_{n}'] = rc1.text_input(f"{mat} (Rough)", d.get(f'{mat} (Rough)_{n}', ''),
                        key=f"{mat_key}_rough_{n}")
                    d[f'{mat} (Face)_{n}'] = rc2.text_input(f"{mat} (Face)", d.get(f'{mat} (Face)_{n}', ''),
                        key=f"{mat_key}_face_{n}")
                for mat in ['Adsa 0', 'Sen 1']:
                    mat_key = mat.replace(' ', '_')
                    d[f'{mat} (Rough)_{n}'] = st.text_input(f"{mat} (Rough Only)", d.get(f'{mat} (Rough)_{n}', ''),
                        key=f"{mat_key}_rough_only_{n}")

            calc = calculate_mix_info(d, entry_num)
            st.write(f"**Calculated Totals (Auto-updates):**")
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Cement 42.5 (Tons)", calc.get(f'Cement 42.5 (Calculated)_{entry_num}', 0))
            m2.metric("Cement 52.5 (Tons)", calc.get(f'Cement 52.5 (Calculated)_{entry_num}', 0))
            m3.metric("White Cement (Tons)", calc.get(f'White Cement (Calculated)_{entry_num}', 0))
            m4.metric("Sand (Kg)", calc.get(f'Sand (Calculated)_{entry_num}', 0))
            m1.metric("Adsa 0 (Kg)", calc.get(f'Adsa 0 (Calculated)_{entry_num}', 0))
            m2.metric("Sen 1 (Kg)", calc.get(f'Sen 1 (Calculated)_{entry_num}', 0))
            m3.metric("Oxide (Kg)", calc.get(f'Oxide (Calculated)_{entry_num}', 0))
            return calc

        st.divider()
        calc1 = render_product_editor(1)
        st.divider()
        calc2 = render_product_editor(2)

        st.divider()
        st.subheader("⚙️ MAINTENANCE & DOWNTIME")
        mc1, mc2 = st.columns(2)
        with mc1:
            st.write("**Electrical Issues**")
            for i in range(4):
                cols = st.columns([1,1,1])
                d["Maintenance Electricity"][i] = cols[0].checkbox(f"Elec #{i+1}", d["Maintenance Electricity"][i])
                d["Maintenance Electricity From"][i] = cols[1].text_input(f"From", d["Maintenance Electricity From"][i], key=f"ef_{i}")
                d["Maintenance Electricity To"][i] = cols[2].text_input(f"To", d["Maintenance Electricity To"][i], key=f"et_{i}")
        with mc2:
            st.write("**Mechanical Issues**")
            for i in range(4):
                cols = st.columns([1,1,1])
                d["Maintenance Mechanical"][i] = cols[0].checkbox(f"Mech #{i+1}", d["Maintenance Mechanical"][i])
                d["Maintenance Mechanical From"][i] = cols[1].text_input(f"From", d["Maintenance Mechanical From"][i], key=f"mf_{i}")
                d["Maintenance Mechanical To"][i] = cols[2].text_input(f"To", d["Maintenance Mechanical To"][i], key=f"mt_{i}")

        st.divider()
        st.subheader("🚜 TOOLS DOWNTIME")
        for tool in ['Loader', 'Clark', 'Solar']:
            st.write(f"**{tool}**")
            for i in range(2):
                cols = st.columns([1, 2, 2])
                d[f"Tools {tool}"][i] = cols[0].checkbox(f"{tool} #{i+1}", d[f"Tools {tool}"][i])
                d[f"Tools {tool} From"][i] = cols[1].text_input("From", d[f"Tools {tool} From"][i], key=f"tl_{tool}_f{i}")
                d[f"Tools {tool} To"][i] = cols[2].text_input("To", d[f"Tools {tool} To"][i], key=f"tl_{tool}_t{i}")

        st.divider()
        st.subheader("⚠️ MISSING MATERIAL")
        for m in ["Cement 42.5", "Cement 52.5", "White Cement", "Sand", "Adsa 0", "Sen 1", "Oxide Red", "Oxide Yellow", "Oxide Black"]:
            cols = st.columns([2, 1, 1])
            d[f"Missing {m}"] = cols[0].checkbox(m, d[f"Missing {m}"])
            d[f"Missing {m} From"] = cols[1].text_input("From", d[f"Missing {m} From"], key=f"miss_{m}_f")
            d[f"Missing {m} To"] = cols[2].text_input("To", d[f"Missing {m} To"], key=f"miss_{m}_t")

        st.divider()
        st.subheader("✍️ SIGN-OFF")
        sc1, sc2, sc3 = st.columns(3)
        d['Supervisor Name'] = sc1.text_input("Supervisor", d.get('Supervisor Name', ''), key="so_supervisor")
        d['Mixer Operator'] = sc1.text_input("Mixer Operator", d.get('Mixer Operator', ''), key="so_mixer")
        d['Machine Operator'] = sc2.text_input("Machine Operator", d.get('Machine Operator', ''), key="so_machine_op")
        d['Forklift Operator'] = sc2.text_input("Forklift Operator", d.get('Forklift Operator', ''), key="so_forklift")
        d['Verified By'] = sc3.text_input("Verified By", d.get('Verified By', ''), key="so_verified")

        # Store the calculations so the export tab has the updated versions
        st.session_state.calc1 = calc1
        st.session_state.calc2 = calc2
    else:
        st.info("👆 Upload and scan documents first.")

with tab3:
    st.header("💾 Export to Excel")
    if st.session_state.extracted_data:
        if st.button("📥 Download Excel File", type="primary"):
            excel_file = export_to_excel(
                st.session_state.extracted_data, 
                st.session_state.get('calc1', {}), 
                st.session_state.get('calc2', {}), 
                st.session_state.combined_image_bytes
            )
            if excel_file:
                with open(excel_file, 'rb') as f:
                    st.download_button(
                        label="⬇️ Click to Download Report", 
                        data=f.read(), 
                        file_name=f"Production_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", 
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
    else:
        st.warning("⚠️ No data available to export.")